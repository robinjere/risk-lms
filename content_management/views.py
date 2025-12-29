from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Avg, Count, Q
from django.conf import settings
from courses.models import Course, Enrollment
from videos.models import Video, VideoSubtitle, VideoProgress, InteractiveCourse, InteractiveCourseProgress
from quizzes.models import Question, QuestionOption, QuizAttempt, QuizAnswer
from accounts.models import User
import json
import os
import re
import tempfile
import zipfile
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging
from django.utils import timezone
from django.utils.dateparse import parse_datetime

logger = logging.getLogger(__name__)

# Try to import moviepy, fallback to basic duration if not available
try:
    from moviepy.editor import VideoFileClip
    MOVIEPY_AVAILABLE = True
except ImportError:
    MOVIEPY_AVAILABLE = False
    print("MoviePy not available, using basic duration estimation")

def calculate_video_duration(video_file):
    """Calculate video duration with better error handling"""
    try:
        # Reset file pointer to beginning
        video_file.seek(0)
        file_size = video_file.size
        
        if MOVIEPY_AVAILABLE:
            try:
                # Use moviepy for accurate duration calculation
                with tempfile.NamedTemporaryFile(delete=False, suffix='.webm') as temp_file:
                    # Write the video content to temp file
                    video_file.seek(0)
                    for chunk in video_file.chunks():
                        temp_file.write(chunk)
                    temp_file_path = temp_file.name
                
                # Calculate duration using moviepy
                video_clip = VideoFileClip(temp_file_path)
                duration = video_clip.duration
                video_clip.close()
                
                # Clean up temp file
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
                
                # Validate duration - MoviePy should give accurate results
                if duration and duration > 0:
                    print(f"MoviePy calculated duration: {duration} seconds")
                    return int(duration)
                else:
                    print(f"MoviePy returned invalid duration: {duration}, using fallback")
                    
            except Exception as moviepy_error:
                print(f"MoviePy error: {moviepy_error}")
        
        # Improved fallback estimation for browser-recorded WebM files
        if file_size > 0:
            # Browser recordings are typically much more compressed
            # WebM from browser MediaRecorder: roughly 100-200 KB per second
            # For a 3:14 video (194 seconds), expect ~19-38 MB file
            
            # More conservative estimation for WebM browser recordings
            bytes_per_second = 150 * 1024  # ~150KB per second average
            estimated_duration = file_size / bytes_per_second
            
            # Clamp to reasonable bounds (10 seconds to 1 hour)
            estimated_duration = max(10, min(3600, estimated_duration))
            
            print(f"Fallback estimation: {file_size} bytes / {bytes_per_second} = {estimated_duration} seconds")
            return int(estimated_duration)
        
        print("No file size available, using default duration")
        return 30  # Default fallback
            
    except Exception as e:
        print(f"Error calculating video duration: {e}")
        return 30  # Safe default

@login_required
def content_dashboard(request):
    """Content management dashboard for Head of Risk and Risk & Compliance Specialist"""
    if not request.user.can_upload_content():
        messages.error(request, "You don't have permission to access this page.")
        return redirect('courses:dashboard')
    
    courses = Course.objects.filter(created_by=request.user).order_by('-created_at')
    total_videos = Video.objects.filter(course__created_by=request.user).count()
    total_questions = Question.objects.filter(course__created_by=request.user).count()
    
    # Get user performance metrics for dashboard display
    user_performance_data = []
    banker_users = User.objects.filter(role='banker')
    
    for user in banker_users:
        # Get user's course enrollments
        user_enrollments = Enrollment.objects.filter(user=user)
        
        if user_enrollments.exists():
            # Calculate overall progress
            total_courses = user_enrollments.count()
            completed_courses = user_enrollments.filter(is_completed=True).count()
            completion_rate = (completed_courses / total_courses * 100) if total_courses > 0 else 0
            
            # Get quiz performance
            user_quiz_attempts = QuizAttempt.objects.filter(
                user=user, completed_at__isnull=False
            )
            
            if user_quiz_attempts.exists():
                best_score = user_quiz_attempts.order_by('-score').first().score
                avg_score = user_quiz_attempts.aggregate(Avg('score'))['score__avg']
                total_attempts = user_quiz_attempts.count()
                
                # Get answer accuracy
                user_answers = QuizAnswer.objects.filter(attempt__user=user)
                total_answers = user_answers.count()
                correct_answers = user_answers.filter(is_correct=True).count()
                accuracy = (correct_answers / total_answers * 100) if total_answers > 0 else 0
                
                # Get latest activity
                latest_attempt = user_quiz_attempts.first()
                
                user_performance_data.append({
                    'user': user,
                    'total_courses': total_courses,
                    'completed_courses': completed_courses,
                    'completion_rate': round(completion_rate, 1),
                    'best_score': round(best_score, 1),
                    'avg_score': round(avg_score, 1),
                    'total_attempts': total_attempts,
                    'accuracy': round(accuracy, 1),
                    'total_answers': total_answers,
                    'correct_answers': correct_answers,
                    'last_activity': latest_attempt.completed_at,
                    'performance_level': 'Excellent' if avg_score >= 85 else 'Good' if avg_score >= 70 else 'Needs Improvement'
                })
    
    # Sort by performance (best score descending)
    user_performance_data.sort(key=lambda x: x['best_score'], reverse=True)
    
    # Calculate summary statistics
    total_bankers = len(user_performance_data)
    excellent_performers = len([u for u in user_performance_data if u['performance_level'] == 'Excellent'])
    good_performers = len([u for u in user_performance_data if u['performance_level'] == 'Good'])
    needs_improvement = len([u for u in user_performance_data if u['performance_level'] == 'Needs Improvement'])
    
    context = {
        'courses': courses,
        'total_videos': total_videos,
        'total_questions': total_questions,
        'user_performance_data': user_performance_data[:10],  # Top 10 for dashboard
        'all_performance_data': user_performance_data,  # All data for detailed view
        'total_bankers': total_bankers,
        'excellent_performers': excellent_performers,
        'good_performers': good_performers,
        'needs_improvement': needs_improvement,
    }
    return render(request, 'content/dashboard.html', context)

@login_required
def upload_course(request):
    """Create or edit course"""
    if not request.user.can_upload_content():
        messages.error(request, "You don't have permission to upload courses.")
        return redirect('courses:dashboard')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        passing_score = request.POST.get('passing_score', 80)
        is_published = request.POST.get('is_published') == 'on'
        
        course = Course.objects.create(
            title=title,
            description=description,
            passing_score=passing_score,
            is_published=is_published,
            created_by=request.user
        )
        
        if 'thumbnail' in request.FILES:
            course.thumbnail = request.FILES['thumbnail']
            course.save()
        
        messages.success(request, f'Course "{title}" created successfully!')
        return redirect('content:video_upload', course_id=course.id)
    
    return render(request, 'content/upload_course.html')

@login_required
def video_upload(request, course_id):
    """Upload or record video with automatic translation"""
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        duration = request.POST.get('duration')
        order_index = request.POST.get('order_index', 0)
        
        # Validate duration
        try:
            duration_seconds = int(duration) if duration else 0
            if duration_seconds <= 0:
                messages.error(request, 'Please provide a valid video duration in seconds.')
                return redirect('content:video_upload', course_id=course.id)
        except ValueError:
            messages.error(request, 'Invalid duration format. Please enter duration in seconds.')
            return redirect('content:video_upload', course_id=course.id)
        
        video = Video.objects.create(
            course=course,
            title=title,
            description=description,
            duration=duration_seconds,
            order_index=int(order_index)
        )
        
        if 'video_file' in request.FILES:
            video_file = request.FILES['video_file']
            video.video_file = video_file
            
            # Try to verify duration from uploaded file if needed
            try:
                calculated_duration = calculate_video_duration(video_file)
                if calculated_duration and abs(calculated_duration - duration_seconds) > 5:
                    # If calculated duration differs significantly from provided duration
                    messages.warning(request, 
                        f'Note: Calculated duration ({calculated_duration}s) differs from provided duration ({duration_seconds}s). '
                        'Using provided duration.')
            except Exception as e:
                # Duration calculation failed, but continue with provided duration
                pass
                
            video.save()
        
        # Handle automatic translation (will be processed in background)
        target_languages = request.POST.getlist('languages')
        if target_languages:
            # Queue translation task (using Celery in production)
            from .tasks import translate_video_audio
            translate_video_audio.delay(video.id, target_languages)
        
        # Success message with duration info
        minutes = duration_seconds // 60
        seconds = duration_seconds % 60
        time_display = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
        
        messages.success(request, 
            f'Video "{title}" uploaded successfully! Duration: {time_display} '
            'Translation in progress...')
        return redirect('content:course_detail', course_id=course.id)
    
    videos = Video.objects.filter(course=course).order_by('order_index')
    context = {
        'course': course,
        'videos': videos,
    }
    return render(request, 'content/video_upload.html', context)

@login_required
@require_POST
def upload_subtitles_view(request, video_id):
    """Upload subtitles for a video (Admin only)"""
    if not request.user.is_risk_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    video = get_object_or_404(Video, id=video_id)
    
    try:
        subtitle_file = request.FILES.get('subtitle_file')
        language_code = request.POST.get('language_code', 'en')
        language_name = request.POST.get('language_name', 'English')
        
        if not subtitle_file:
            return JsonResponse({'error': 'No subtitle file provided'}, status=400)
        
        # Check if subtitle already exists for this language
        existing_subtitle = VideoSubtitle.objects.filter(
            video=video, 
            language_code=language_code
        ).first()
        
        if existing_subtitle:
            # Update existing subtitle
            existing_subtitle.subtitle_file = subtitle_file
            existing_subtitle.language_name = language_name
            existing_subtitle.save()
            subtitle = existing_subtitle
        else:
            # Create new subtitle
            subtitle = VideoSubtitle.objects.create(
                video=video,
                language_code=language_code,
                language_name=language_name,
                subtitle_file=subtitle_file
            )
        
        return JsonResponse({
            'success': True,
            'subtitle_id': subtitle.id,
            'language_name': subtitle.language_name,
            'language_code': subtitle.language_code,
            'message': f'Subtitle for {language_name} uploaded successfully'
        })
    
    except Exception as e:
        return JsonResponse({
            'error': f'Failed to upload subtitle: {str(e)}'
        }, status=500)

@login_required
@require_POST
def save_recorded_video(request):
    """Save video recorded from browser"""
    if not request.user.can_upload_content():
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        course_id = request.POST.get('course_id')
        title = request.POST.get('title')
        video_blob = request.FILES.get('video_blob')
        
        if not video_blob:
            return JsonResponse({'error': 'No video file provided'}, status=400)
        
        course = get_object_or_404(Course, id=course_id, created_by=request.user)
        
        # Calculate video duration
        duration = calculate_video_duration(video_blob)
        
        # Reset file pointer after duration calculation
        video_blob.seek(0)
        
        video = Video.objects.create(
            course=course,
            title=title,
            video_file=video_blob,
            duration=duration,
            file_size=video_blob.size,
            order_index=Video.objects.filter(course=course).count()
        )
        
        return JsonResponse({
            'success': True,
            'video_id': video.id,
            'duration': duration,
            'message': f'Video recorded successfully! Duration: {duration//60}:{duration%60:02d}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Failed to save video: {str(e)}'
        }, status=500)


def parse_captivate_metadata(extracted_path):
    """Parse metadata from Adobe Captivate course package"""
    metadata = {
        'title': '',
        'duration_minutes': 0,
        'total_slides': 0,
        'resolution_width': 1280,
        'resolution_height': 720,
        'entry_file': 'index.html',
        'content_type': 'captivate'
    }
    
    # Look for CPM.js or similar config files
    for root, dirs, files in os.walk(extracted_path):
        for file in files:
            file_path = os.path.join(root, file)
            
            # Check CPM.js for Captivate metadata
            if file == 'CPM.js':
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        
                        # Extract project title
                        title_match = re.search(r'projectTitle\s*[=:]\s*["\']([^"\']+)["\']', content)
                        if title_match:
                            metadata['title'] = title_match.group(1)
                        
                        # Extract duration
                        duration_match = re.search(r'projectDuration\s*[=:]\s*(\d+)', content)
                        if duration_match:
                            # Duration is usually in milliseconds or seconds
                            duration = int(duration_match.group(1))
                            if duration > 10000:  # Likely milliseconds
                                metadata['duration_minutes'] = duration // 60000
                            else:
                                metadata['duration_minutes'] = duration // 60
                        
                        # Extract slide count
                        slide_match = re.search(r'totalSlides\s*[=:]\s*(\d+)', content)
                        if slide_match:
                            metadata['total_slides'] = int(slide_match.group(1))
                        
                        # Extract resolution
                        width_match = re.search(r'stageWidth\s*[=:]\s*(\d+)', content)
                        height_match = re.search(r'stageHeight\s*[=:]\s*(\d+)', content)
                        if width_match:
                            metadata['resolution_width'] = int(width_match.group(1))
                        if height_match:
                            metadata['resolution_height'] = int(height_match.group(1))
                            
                except Exception as e:
                    print(f"Error parsing CPM.js: {e}")
            
            # Check project.txt for additional metadata (JSON format for Captivate)
            elif file == 'project.txt':
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        
                        # Try to parse as JSON (Captivate 11+ format)
                        try:
                            project_data = json.loads(content)
                            
                            # Extract metadata from JSON structure
                            if 'metadata' in project_data:
                                meta = project_data['metadata']
                                
                                # Title
                                if meta.get('title') and not metadata['title']:
                                    metadata['title'] = meta['title']
                                
                                # Total slides
                                if meta.get('totalSlides'):
                                    metadata['total_slides'] = int(meta['totalSlides'])
                                
                                # Duration from frames
                                if meta.get('durationInFrames') and meta.get('frameRate'):
                                    duration_seconds = meta['durationInFrames'] / meta['frameRate']
                                    metadata['duration_minutes'] = int(duration_seconds / 60)
                                
                                # Resolution
                                if meta.get('width'):
                                    metadata['resolution_width'] = int(meta['width'])
                                if meta.get('height'):
                                    metadata['resolution_height'] = int(meta['height'])
                                
                                # Entry file
                                if meta.get('launchFile'):
                                    metadata['entry_file'] = meta['launchFile']
                            
                            # Also check for title in contentStructure (project name)
                            if 'contentStructure' in project_data:
                                for item in project_data['contentStructure']:
                                    if item.get('class') == 'project' and item.get('title'):
                                        if not metadata['title']:
                                            metadata['title'] = item['title']
                                        break
                                        
                        except json.JSONDecodeError:
                            # Fallback to regex for non-JSON format
                            if not metadata['title']:
                                title_match = re.search(r'name\s*[=:]\s*(.+)', content)
                                if title_match:
                                    metadata['title'] = title_match.group(1).strip()
                                    
                            if not metadata['duration_minutes']:
                                duration_match = re.search(r'duration\s*[=:]\s*(\d+)', content)
                                if duration_match:
                                    metadata['duration_minutes'] = int(duration_match.group(1))
                                
                except Exception as e:
                    print(f"Error parsing project.txt: {e}")
            
            # Check index.html for title
            elif file == 'index.html':
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        
                        if not metadata['title']:
                            title_match = re.search(r'<title>([^<]+)</title>', content, re.IGNORECASE)
                            if title_match:
                                metadata['title'] = title_match.group(1).strip()
                                
                except Exception as e:
                    print(f"Error parsing index.html: {e}")
    
    return metadata


@login_required
def interactive_course_list(request):
    """List all interactive courses for browsing and enrollment"""
    from videos.models import InteractiveCourse, InteractiveCourseProgress
    from courses.models import Enrollment
    from quizzes.models import Question
    
    # Get all interactive courses with their parent courses
    interactive_courses = InteractiveCourse.objects.filter(
        is_active=True,
        course__is_published=True
    ).select_related('course', 'created_by').order_by('-created_at')
    
    # Get user's enrolled courses
    user_enrollments = set()
    user_progress = {}
    
    if request.user.is_authenticated:
        user_enrollments = set(
            Enrollment.objects.filter(user=request.user).values_list('course_id', flat=True)
        )
        
        # Get progress for each interactive course
        progress_records = InteractiveCourseProgress.objects.filter(
            user=request.user,
            interactive_course__in=interactive_courses
        )
        for p in progress_records:
            user_progress[p.interactive_course_id] = p
    
    # Get question counts for each interactive course
    question_counts = {}
    for ic in interactive_courses:
        question_counts[ic.id] = Question.objects.filter(interactive_course=ic).count()
    
    # Prepare data for template
    courses_data = []
    for ic in interactive_courses:
        is_enrolled = ic.course_id in user_enrollments
        progress = user_progress.get(ic.id)
        
        courses_data.append({
            'interactive_course': ic,
            'course': ic.course,
            'is_enrolled': is_enrolled,
            'progress': progress,
            'can_access': is_enrolled or request.user.is_risk_admin(),
            'question_count': question_counts.get(ic.id, 0),
        })
    
    # Get available courses for admin to add new interactive content
    available_courses = []
    is_admin = request.user.is_risk_admin() if hasattr(request.user, 'is_risk_admin') else False
    if is_admin:
        available_courses = Course.objects.filter(is_published=True).order_by('title')
    
    context = {
        'courses_data': courses_data,
        'total_modules': interactive_courses.count(),
        'is_admin': is_admin,
        'available_courses': available_courses,
    }
    
    return render(request, 'content/interactive_list.html', context)


@login_required
def upload_interactive_course_new(request):
    """Upload Adobe Captivate/SCORM package from interactive list page (without course_id in URL)"""
    # Only risk admins can upload
    if not request.user.is_risk_admin():
        messages.error(request, 'You do not have permission to upload interactive courses.')
        return redirect('content:interactive_list')
    
    if request.method != 'POST':
        return redirect('content:interactive_list')
    
    # Check if creating a new course or using existing
    course_option = request.POST.get('course_option', 'existing')
    course = None
    
    if course_option == 'new':
        # Create a new course
        new_course_title = request.POST.get('new_course_title', '').strip()
        new_course_description = request.POST.get('new_course_description', '').strip()
        
        if not new_course_title:
            messages.error(request, 'Please enter a title for the new course.')
            return redirect('content:interactive_list')
        
        # Create the new course
        course = Course.objects.create(
            title=new_course_title,
            description=new_course_description or f'Interactive course: {new_course_title}',
            created_by=request.user,
            is_published=True,
            passing_score=80
        )
        messages.info(request, f'New course "{course.title}" created.')
    else:
        # Use existing course
        course_id = request.POST.get('course_id')
        if not course_id:
            messages.error(request, 'Please select a course or create a new one.')
            return redirect('content:interactive_list')
        
        course = get_object_or_404(Course, id=course_id)
    
    title = request.POST.get('title', '')
    description = request.POST.get('description', '')
    content_type = request.POST.get('content_type', 'captivate')
    order_index = request.POST.get('order_index', 0)
    
    package_file = request.FILES.get('package_file')
    
    if not package_file:
        messages.error(request, 'Please select a package file to upload.')
        return redirect('content:interactive_list')
    
    # Validate file is a zip
    if not package_file.name.endswith('.zip'):
        messages.error(request, 'Package must be a ZIP file.')
        return redirect('content:interactive_list')
    
    try:
        # Create a unique extraction path
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_title = re.sub(r'[^\w\-]', '_', title or 'course')[:50]
        extract_folder = f'interactive_courses/{timestamp}_{safe_title}'
        extract_path = os.path.join(settings.MEDIA_ROOT, extract_folder)
        
        # Create directory
        os.makedirs(extract_path, exist_ok=True)
        
        # Save and extract the zip file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as temp_zip:
            for chunk in package_file.chunks():
                temp_zip.write(chunk)
            temp_zip_path = temp_zip.name
        
        # Extract the zip file
        with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_path)
        
        # Clean up temp file
        os.unlink(temp_zip_path)
        
        # Check if files were extracted to a subdirectory
        extracted_items = os.listdir(extract_path)
        if len(extracted_items) == 1 and os.path.isdir(os.path.join(extract_path, extracted_items[0])):
            # Move contents up one level
            sub_dir = os.path.join(extract_path, extracted_items[0])
            for item in os.listdir(sub_dir):
                shutil.move(os.path.join(sub_dir, item), extract_path)
            os.rmdir(sub_dir)
        
        # Parse metadata from extracted files
        metadata = parse_captivate_metadata(extract_path)
        
        # Find entry file
        entry_file = 'index.html'
        for possible_entry in ['index.html', 'index.htm', 'default.html', 'story.html']:
            if os.path.exists(os.path.join(extract_path, possible_entry)):
                entry_file = possible_entry
                break
        
        # Safe integer conversion helper
        def safe_int_convert(value, default=0):
            if value is None or value == '':
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default
        
        # Create the interactive course record
        interactive_course = InteractiveCourse.objects.create(
            course=course,
            title=title or metadata.get('title', 'Untitled Course'),
            description=description or metadata.get('description', ''),
            content_type=content_type,
            package_file=package_file,
            extracted_path=extract_folder,
            entry_file=entry_file,
            duration_minutes=safe_int_convert(metadata.get('duration_minutes')),
            total_slides=safe_int_convert(metadata.get('total_slides')),
            resolution_width=safe_int_convert(metadata.get('resolution_width'), 1280),
            resolution_height=safe_int_convert(metadata.get('resolution_height'), 720),
            order_index=safe_int_convert(order_index),
            created_by=request.user
        )
        
        messages.success(request, f'Interactive course "{interactive_course.title}" uploaded successfully!')
        return redirect('content:interactive_list')
        
    except Exception as e:
        messages.error(request, f'Error uploading package: {str(e)}')
        return redirect('content:interactive_list')


@login_required
def upload_interactive_course(request, course_id):
    """Upload Adobe Captivate/SCORM package"""
    # Allow access if user is risk admin or created the course
    if request.user.is_risk_admin():
        course = get_object_or_404(Course, id=course_id)
    else:
        course = get_object_or_404(Course, id=course_id, created_by=request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title', '')
        description = request.POST.get('description', '')
        content_type = request.POST.get('content_type', 'captivate')
        order_index = request.POST.get('order_index', 0)
        
        package_file = request.FILES.get('package_file')
        
        if not package_file:
            messages.error(request, 'Please select a package file to upload.')
            return redirect('content:upload_interactive', course_id=course.id)
        
        # Validate file is a zip
        if not package_file.name.endswith('.zip'):
            messages.error(request, 'Package must be a ZIP file.')
            return redirect('content:upload_interactive', course_id=course.id)
        
        try:
            # Create a unique extraction path
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_title = re.sub(r'[^\w\-]', '_', title or 'course')[:50]
            extract_folder = f'interactive_courses/{timestamp}_{safe_title}'
            extract_path = os.path.join(settings.MEDIA_ROOT, extract_folder)
            
            # Create directory
            os.makedirs(extract_path, exist_ok=True)
            
            # Save and extract the zip file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as temp_zip:
                for chunk in package_file.chunks():
                    temp_zip.write(chunk)
                temp_zip_path = temp_zip.name
            
            # Extract the zip file
            with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_path)
            
            # Clean up temp file
            os.unlink(temp_zip_path)
            
            # Check if files were extracted to a subdirectory
            extracted_items = os.listdir(extract_path)
            if len(extracted_items) == 1 and os.path.isdir(os.path.join(extract_path, extracted_items[0])):
                # Move contents up one level
                sub_dir = os.path.join(extract_path, extracted_items[0])
                for item in os.listdir(sub_dir):
                    shutil.move(os.path.join(sub_dir, item), extract_path)
                os.rmdir(sub_dir)
            
            # Parse metadata from extracted files
            metadata = parse_captivate_metadata(extract_path)
            
            # Use metadata title if no title provided
            if not title and metadata['title']:
                title = metadata['title']
            
            # Find entry file
            entry_file = 'index.html'
            if os.path.exists(os.path.join(extract_path, 'index.html')):
                entry_file = 'index.html'
            elif os.path.exists(os.path.join(extract_path, 'index.htm')):
                entry_file = 'index.htm'
            
            # Reset package file pointer and save
            package_file.seek(0)
            
            # Safely parse integer values from form
            def safe_int(value, default=0):
                try:
                    return int(value) if value else default
                except (ValueError, TypeError):
                    return default
            
            # Get duration and slides - prefer metadata, fallback to form input
            final_duration = metadata['duration_minutes'] or safe_int(request.POST.get('duration_minutes'))
            final_slides = metadata['total_slides'] or safe_int(request.POST.get('total_slides'))
            final_order = safe_int(order_index)
            
            # Create the interactive course record
            interactive_course = InteractiveCourse.objects.create(
                course=course,
                title=title or metadata['title'] or 'Untitled Interactive Course',
                description=description,
                content_type=content_type,
                package_file=package_file,
                extracted_path=extract_folder,
                entry_file=entry_file,
                duration_minutes=final_duration,
                total_slides=final_slides,
                resolution_width=metadata['resolution_width'],
                resolution_height=metadata['resolution_height'],
                order_index=final_order,
                created_by=request.user
            )
            
            # Handle thumbnail upload
            if 'thumbnail' in request.FILES:
                interactive_course.thumbnail = request.FILES['thumbnail']
                interactive_course.save()
            
            # Use the final title for the success message
            final_title = interactive_course.title
            messages.success(request, 
                f'Interactive course "{final_title}" uploaded successfully! '
                f'Duration: {final_duration} minutes, '
                f'Slides: {final_slides}'
            )
            return redirect('content:course_detail', course_id=course.id)
            
        except zipfile.BadZipFile:
            messages.error(request, 'Invalid ZIP file. Please upload a valid package.')
            return redirect('content:upload_interactive', course_id=course.id)
        except Exception as e:
            messages.error(request, f'Error uploading package: {str(e)}')
            return redirect('content:upload_interactive', course_id=course.id)
    
    # GET request - show upload form
    interactive_courses = InteractiveCourse.objects.filter(course=course).order_by('order_index')
    context = {
        'course': course,
        'interactive_courses': interactive_courses,
    }
    return render(request, 'content/upload_interactive.html', context)


@login_required
def play_interactive_course(request, course_id, interactive_id):
    """Play/launch an interactive course"""
    course = get_object_or_404(Course, id=course_id)
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id, course=course)
    
    # Check if user is enrolled
    if not request.user.is_risk_admin():
        enrollment = Enrollment.objects.filter(user=request.user, course=course).first()
        if not enrollment:
            messages.error(request, 'You must be enrolled in this course to view content.')
            return redirect('courses:course_detail', course_id=course.id)
    
    # Get or create progress record
    progress, created = InteractiveCourseProgress.objects.get_or_create(
        user=request.user,
        interactive_course=interactive_course
    )

    # Resume safety: never resume beyond the highest legitimately reached slide.
    if progress.highest_slide_reached and progress.current_slide > progress.highest_slide_reached:
        progress.current_slide = progress.highest_slide_reached
        progress.save(update_fields=['current_slide', 'updated_at'])
    
    # Calculate progress dash offset for circular SVG (circumference is 339.292)
    circumference = 339.292
    progress_dashoffset = circumference - (progress.completion_percentage / 100 * circumference)
    
    # Create range of slides for template
    slide_range = range(1, interactive_course.total_slides + 1)
    
    context = {
        'course': course,
        'interactive_course': interactive_course,
        'progress': progress,
        'launch_url': interactive_course.get_launch_url(),
        'slide_range': slide_range,
        'progress_dashoffset': progress_dashoffset,
    }
    return render(request, 'content/play_interactive.html', context)


@login_required
@require_POST
def update_interactive_progress(request, interactive_id):
    """Update progress for interactive course - tracks slides, quiz, and completion"""
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    
    try:
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {
                    'success': False,
                    'error': 'Invalid JSON payload.',
                    'code': 'invalid_json',
                },
                status=400,
            )
        
        progress, created = InteractiveCourseProgress.objects.get_or_create(
            user=request.user,
            interactive_course=interactive_course
        )
        
        total_slides = interactive_course.total_slides
        min_time_per_slide_seconds = progress.get_min_time_per_slide_seconds()
        now = timezone.now()

        def _parse_int(value, field_name):
            try:
                return int(value)
            except (TypeError, ValueError):
                raise ValueError(f'Invalid {field_name}. Must be an integer.')

        def _parse_ts(value):
            if not value:
                return None
            dt = parse_datetime(str(value))
            if not dt:
                return None
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt, timezone.get_current_timezone())
            return dt

        def _reject(message, *, status=400, code='invalid_request'):
            return JsonResponse(
                {
                    'success': False,
                    'error': message,
                    'code': code,
                    'current_slide': progress.current_slide,
                    'highest_slide_reached': progress.highest_slide_reached,
                    'allowed_next_slide': (progress.highest_slide_reached or 0) + 1,
                    'total_slides': total_slides,
                },
                status=status,
            )

        try:
            # Reject obvious slide-jump attempts early (even if client sends highest_slide_reached).
            if 'highest_slide_reached' in data:
                requested_highest = _parse_int(data.get('highest_slide_reached'), 'highest_slide_reached')
                if requested_highest > (progress.highest_slide_reached or 0) + 1:
                    progress.skip_attempts = (progress.skip_attempts or 0) + 1
                    progress.save(update_fields=['skip_attempts', 'updated_at'])
                    logger.warning(
                        'Interactive skip attempt: user=%s course=%s requested_highest=%s current=%s highest=%s',
                        request.user.id,
                        interactive_course.id,
                        requested_highest,
                        progress.current_slide,
                        progress.highest_slide_reached,
                    )
                    return _reject(
                        'Cannot unlock future slides. Complete the previous slide to continue.',
                        status=403,
                        code='slide_skip',
                    )

            requested_current_slide = None
            if 'current_slide' in data:
                requested_current_slide = _parse_int(data.get('current_slide'), 'current_slide')
        except ValueError as ve:
            return _reject(str(ve), status=400, code='invalid_payload')

        # Slide completion must be explicit (Next button). Enforce sequential completion + minimum time.
        slide_completed = None
        if 'slide_completed' in data:
            try:
                slide_completed = _parse_int(data.get('slide_completed'), 'slide_completed')
            except ValueError as ve:
                return _reject(str(ve), status=400, code='invalid_payload')
            if slide_completed < 1 or (total_slides and slide_completed > total_slides):
                return _reject('Invalid slide number.', status=400, code='invalid_slide')

            if slide_completed > (progress.highest_slide_reached or 0) + 1:
                progress.skip_attempts = (progress.skip_attempts or 0) + 1
                progress.save(update_fields=['skip_attempts', 'updated_at'])
                logger.warning(
                    'Interactive skip attempt (complete): user=%s course=%s slide_completed=%s current=%s highest=%s',
                    request.user.id,
                    interactive_course.id,
                    slide_completed,
                    progress.current_slide,
                    progress.highest_slide_reached,
                )
                return _reject(
                    f'Cannot complete slide {slide_completed} yet. Complete previous slides first.',
                    status=403,
                    code='slide_skip',
                )

            # Ensure we have a recorded start time for this slide.
            progress.start_slide(slide_completed)
            started_at = _parse_ts(progress.slide_started_at.get(str(slide_completed)))
            if not started_at:
                started_at = now
                progress.slide_started_at[str(slide_completed)] = started_at.isoformat()

            elapsed = (now - started_at).total_seconds()
            if elapsed < min_time_per_slide_seconds:
                remaining = int(max(1, min_time_per_slide_seconds - elapsed))
                progress.save(update_fields=['slide_started_at', 'updated_at'])
                return _reject(
                    f'Slide unlocks in {remaining}s. Please finish the slide before continuing.',
                    status=400,
                    code='min_time_not_met',
                )

            progress.mark_slide_completed(slide_completed)

        # Track current slide (allowed only for current/previous/next-unlocked slide).
        if requested_current_slide and requested_current_slide > 0:
            if not progress.can_access_slide(requested_current_slide):
                progress.skip_attempts = (progress.skip_attempts or 0) + 1
                progress.save(update_fields=['skip_attempts', 'updated_at'])
                logger.warning(
                    'Interactive skip attempt: user=%s course=%s requested_current=%s current=%s highest=%s',
                    request.user.id,
                    interactive_course.id,
                    requested_current_slide,
                    progress.current_slide,
                    progress.highest_slide_reached,
                )
                return _reject(
                    f'Slide {requested_current_slide} is locked. Complete previous slides to unlock.',
                    status=403,
                    code='slide_locked',
                )

            progress.current_slide = requested_current_slide
            progress.start_slide(requested_current_slide)
        
        # Track time spent (in minutes)
        if 'time_spent' in data:
            progress.total_time_spent += int(data['time_spent'])
        
        # Handle quiz results from Captivate or LMS quiz
        if 'quiz_score' in data:
            progress.quiz_score = float(data['quiz_score'])
            progress.quiz_attempts += 1
            progress.quiz_passed = progress.quiz_score >= 80  # 80% passing threshold
            
            # If content is completed and quiz passed, mark course as fully completed
            if progress.content_completed and progress.quiz_passed:
                if not progress.is_completed:
                    progress.is_completed = True
                    progress.completed_at = datetime.now()
        
        # Store SCORM data
        if 'scorm_data' in data:
            progress.scorm_data.update(data['scorm_data'])
        
        # Store SCORM suspend_data (for course resume)
        if 'scorm_suspend_data' in data:
            progress.scorm_suspend_data = data['scorm_suspend_data']
        
        # Calculate completion percentage based on highest slide reached
        if total_slides > 0:
            progress.completion_percentage = min(100, int((progress.highest_slide_reached / total_slides) * 100))
        
        # Check content completion (all slides reached OR frontend signals completion)
        if total_slides > 0 and progress.highest_slide_reached >= total_slides:
            if not progress.content_completed:
                progress.content_completed = True
                from django.utils import timezone
                progress.content_completed_at = timezone.now()
        
        # Also accept content_completed from frontend (fallback)
        if data.get('content_completed') == True and not progress.content_completed:
            if progress.highest_slide_reached >= total_slides:
                progress.content_completed = True
                from django.utils import timezone
                progress.content_completed_at = timezone.now()
        
        progress.save()
        
        return JsonResponse({
            'success': True,
            'current_slide': progress.current_slide,
            'highest_slide_reached': progress.highest_slide_reached,
            'completion_percentage': progress.completion_percentage,
            'content_completed': progress.content_completed,
            'quiz_passed': progress.quiz_passed,
            'quiz_score': progress.quiz_score,
            'is_completed': progress.is_completed,
            'can_take_quiz': progress.content_completed,
            'slides_completed': progress.highest_slide_reached,
            'total_slides': total_slides,
            'total_time_spent': progress.total_time_spent
        })
        
    except Exception as e:
        logger.exception('Failed to update interactive progress: user=%s course=%s', request.user.id, interactive_course.id)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def question_bank(request, course_id):
    """Manage question bank for course"""
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    
    if request.method == 'POST':
        question_text = request.POST.get('question_text')
        question_type = request.POST.get('question_type')
        topic = request.POST.get('topic')
        difficulty = request.POST.get('difficulty')
        points = request.POST.get('points', 1)
        
        question = Question.objects.create(
            course=course,
            question_text=question_text,
            question_type=question_type,
            topic=topic,
            difficulty=difficulty,
            points=int(points)
        )
        
        # Add options
        options_data = json.loads(request.POST.get('options', '[]'))
        for idx, opt in enumerate(options_data):
            QuestionOption.objects.create(
                question=question,
                option_text=opt['text'],
                is_correct=opt['is_correct'],
                order_index=idx
            )
        
        messages.success(request, 'Question added successfully!')
        return redirect('content:question_bank', course_id=course.id)
    
    questions = Question.objects.filter(course=course).order_by('topic', '-created_at')
    context = {
        'course': course,
        'questions': questions,
    }
    return render(request, 'content/question_bank.html', context)

@login_required
def course_detail(request, course_id):
    """View course details with videos and questions"""
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    videos = Video.objects.filter(course=course).order_by('order_index')
    questions = Question.objects.filter(course=course).order_by('topic')
    interactive_courses = InteractiveCourse.objects.filter(course=course).order_by('order_index')
    
    # Calculate total duration
    total_duration = sum(video.duration for video in videos)
    
    context = {
        'course': course,
        'videos': videos,
        'questions': questions,
        'interactive_courses': interactive_courses,
        'total_duration': total_duration,
        'can_manage': request.user.is_risk_admin(),
    }
    return render(request, 'content/course_detail.html', context)

@login_required
@require_http_methods(["POST"])
def delete_video(request, video_id):
    """Delete a video (Risk Admin/Compliance only)"""
    if not request.user.is_risk_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    video = get_object_or_404(Video, id=video_id)
    
    # Check if user created this video's course
    if video.course.created_by != request.user:
        return JsonResponse({'error': 'You can only delete videos from courses you created'}, status=403)
    
    try:
        with transaction.atomic():
            # Delete related progress records
            VideoProgress.objects.filter(video=video).delete()
            
            # Delete video file from storage
            if video.video_file:
                try:
                    video.video_file.delete(save=False)
                except:
                    pass  # File might already be deleted
            
            # Delete subtitle files
            for subtitle in video.subtitles.all():
                if subtitle.subtitle_file:
                    try:
                        subtitle.subtitle_file.delete(save=False)
                    except:
                        pass
            
            video_title = video.title
            course_id = video.course.id
            video.delete()
            
            messages.success(request, f'Video "{video_title}" deleted successfully!')
            
        return JsonResponse({
            'success': True,
            'message': f'Video "{video_title}" deleted successfully!',
            'redirect_url': f'/content/course/{course_id}/'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Failed to delete video: {str(e)}'}, status=500)

@login_required
@require_http_methods(["POST"])
def delete_question(request, question_id):
    """Delete a question (Risk Admin/Compliance only)"""
    if not request.user.is_risk_admin():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    question = get_object_or_404(Question, id=question_id)
    
    # Check if user created this question's course
    if question.course.created_by != request.user:
        return JsonResponse({'error': 'You can only delete questions from courses you created'}, status=403)
    
    try:
        with transaction.atomic():
            question_text = question.question_text[:50] + "..." if len(question.question_text) > 50 else question.question_text
            course_id = question.course.id
            question.delete()
            
            messages.success(request, f'Question "{question_text}" deleted successfully!')
            
        return JsonResponse({
            'success': True,
            'message': f'Question deleted successfully!',
            'redirect_url': f'/content/course/{course_id}/questions/'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Failed to delete question: {str(e)}'}, status=500)

@login_required
@require_http_methods(["GET", "POST"])
def edit_course_settings(request, course_id):
    """Edit course settings including time limits (Risk Admin/Compliance only)"""
    if not request.user.is_risk_admin():
        messages.error(request, 'Unauthorized access.')
        return redirect('content:dashboard')
    
    course = get_object_or_404(Course, id=course_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            # Update basic course info
            course.title = request.POST.get('title', course.title)
            course.description = request.POST.get('description', course.description)
            course.passing_score = int(request.POST.get('passing_score', course.passing_score))
            course.is_published = request.POST.get('is_published') == 'on'
            
            # Update completion time settings
            course.completion_time_enabled = request.POST.get('completion_time_enabled') == 'on'
            if course.completion_time_enabled:
                time_limit = request.POST.get('completion_time_limit', '0')
                course.completion_time_limit = max(1, int(time_limit))
            else:
                course.completion_time_limit = 0
            
            course.save()
            
            # Update existing enrollments with new deadlines if time limit changed
            if course.completion_time_enabled and course.completion_time_limit > 0:
                enrollments = Enrollment.objects.filter(course=course, is_completed=False)
                for enrollment in enrollments:
                    if not enrollment.deadline:  # Only set if not already set
                        from datetime import timedelta
                        enrollment.deadline = enrollment.enrolled_at + timedelta(days=course.completion_time_limit)
                        enrollment.save()
            
            messages.success(request, f'Course "{course.title}" updated successfully!')
            return redirect('content:course_detail', course_id=course.id)
            
        except ValueError as e:
            messages.error(request, f'Invalid input: {e}')
        except Exception as e:
            messages.error(request, f'Error updating course: {e}')
    
    context = {
        'course': course,
        'enrollments_count': course.enrollments.count(),
        'active_enrollments': course.enrollments.filter(is_completed=False).count(),
        'completed_enrollments': course.enrollments.filter(is_completed=True).count(),
    }
    
    # Calculate completion rate
    if context['enrollments_count'] > 0:
        context['completion_rate'] = int((context['completed_enrollments'] / context['enrollments_count']) * 100)
    else:
        context['completion_rate'] = 0
    
    return render(request, 'content/edit_course.html', context)


@login_required
def download_course_enrollment_report(request):
    """Download Excel report of course enrollments and completion statistics"""
    if not request.user.is_risk_admin():
        messages.error(request, 'Unauthorized access.')
        return redirect('content:dashboard')
    
    from videos.models import InteractiveCourse, InteractiveCourseProgress
    from certificates.models import Certificate
    
    # Create workbook
    wb = Workbook()
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002B5C", end_color="002B5C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Success styling (green)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    success_font = Font(color="006100")
    
    # Warning styling (yellow)
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    warning_font = Font(color="9C5700")
    
    # Danger styling (red)
    danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    danger_font = Font(color="9C0006")
    
    # Info styling (blue)
    info_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    info_font = Font(color="1F4E79")
    
    # ==========================================
    # Sheet 1: Summary Statistics
    # ==========================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Get all bankers
    all_bankers = User.objects.filter(role='banker')
    total_bankers = all_bankers.count()
    
    # Get interactive courses
    interactive_courses = InteractiveCourse.objects.filter(is_active=True)
    
    # Calculate statistics
    total_logged_in = all_bankers.filter(last_login__isnull=False).count()
    never_logged_in = total_bankers - total_logged_in
    total_certificates = Certificate.objects.filter(is_valid=True).count()
    
    # Summary headers
    ws_summary.cell(row=1, column=1, value="COURSE ENROLLMENT & COMPLETION REPORT").font = Font(bold=True, size=16)
    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}").font = Font(italic=True)
    ws_summary.cell(row=3, column=1, value="Co-operative Bank of Tanzania PLC - Risk Department LMS").font = Font(italic=True)
    
    ws_summary.cell(row=5, column=1, value="OVERALL STATISTICS").font = Font(bold=True, size=12)
    
    stats = [
        ("Total Staff (AD Users)", total_bankers),
        ("Staff Who Have Logged In", total_logged_in),
        ("Staff Never Logged In", never_logged_in),
        ("Total Certificates Issued", total_certificates),
        ("Compliance Rate", f"{round((total_certificates / total_bankers * 100) if total_bankers > 0 else 0, 1)}%"),
    ]
    
    for i, (label, value) in enumerate(stats, 6):
        ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)
    
    # Per-course statistics
    row = len(stats) + 8
    ws_summary.cell(row=row, column=1, value="PER-COURSE STATISTICS").font = Font(bold=True, size=12)
    row += 1
    
    course_headers = ["Course Name", "Total Staff", "Enrolled", "Not Started", "In Progress", "Completed Content", "Certified", "Completion Rate"]
    for col, header in enumerate(course_headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    row += 1
    for ic in interactive_courses:
        enrolled = InteractiveCourseProgress.objects.filter(interactive_course=ic).count()
        not_started = total_bankers - enrolled
        completed_content = InteractiveCourseProgress.objects.filter(interactive_course=ic, content_completed=True).count()
        in_progress = enrolled - completed_content
        certified = Certificate.objects.filter(interactive_course=ic, is_valid=True).count()
        completion_rate = round((certified / enrolled * 100) if enrolled > 0 else 0, 1)
        
        ws_summary.cell(row=row, column=1, value=ic.title)
        ws_summary.cell(row=row, column=2, value=total_bankers)
        ws_summary.cell(row=row, column=3, value=enrolled)
        ws_summary.cell(row=row, column=4, value=not_started)
        ws_summary.cell(row=row, column=5, value=in_progress)
        ws_summary.cell(row=row, column=6, value=completed_content)
        ws_summary.cell(row=row, column=7, value=certified)
        ws_summary.cell(row=row, column=8, value=f"{completion_rate}%")
        row += 1
    
    # ==========================================
    # Sheet 2: Detailed Staff Progress (AD Users)
    # ==========================================
    ws_staff = wb.create_sheet("Staff Progress (AD)")
    
    # Build headers dynamically based on courses
    staff_headers = ["#", "Full Name", "Email (AD Account)", "Last Login", "Login Status"]
    for ic in interactive_courses:
        staff_headers.append(f"{ic.title[:30]} - Status")
        staff_headers.append(f"{ic.title[:30]} - Progress %")
        staff_headers.append(f"{ic.title[:30]} - Started")
        staff_headers.append(f"{ic.title[:30]} - Certificate")
    
    for col, header in enumerate(staff_headers, 1):
        cell = ws_staff.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add staff data
    row = 2
    for idx, banker in enumerate(all_bankers.order_by('first_name', 'last_name'), 1):
        col = 1
        
        # Basic info
        ws_staff.cell(row=row, column=col, value=idx)
        col += 1
        
        ws_staff.cell(row=row, column=col, value=banker.get_full_name() or banker.username)
        col += 1
        
        ws_staff.cell(row=row, column=col, value=banker.email)
        col += 1
        
        # Last login
        if banker.last_login:
            ws_staff.cell(row=row, column=col, value=banker.last_login.strftime('%d %b %Y %H:%M'))
        else:
            ws_staff.cell(row=row, column=col, value="Never")
        col += 1
        
        # Login status with color coding
        login_cell = ws_staff.cell(row=row, column=col)
        if banker.last_login:
            login_cell.value = "Logged In"
            login_cell.fill = success_fill
            login_cell.font = success_font
        else:
            login_cell.value = "Never Logged In"
            login_cell.fill = danger_fill
            login_cell.font = danger_font
        col += 1
        
        # Course-specific data
        for ic in interactive_courses:
            progress = InteractiveCourseProgress.objects.filter(user=banker, interactive_course=ic).first()
            certificate = Certificate.objects.filter(user=banker, interactive_course=ic, is_valid=True).first()
            
            # Status
            status_cell = ws_staff.cell(row=row, column=col)
            if certificate:
                status_cell.value = "Certified"
                status_cell.fill = success_fill
                status_cell.font = success_font
            elif progress and progress.content_completed:
                status_cell.value = "Awaiting Quiz"
                status_cell.fill = info_fill
                status_cell.font = info_font
            elif progress:
                status_cell.value = "In Progress"
                status_cell.fill = warning_fill
                status_cell.font = warning_font
            else:
                status_cell.value = "Not Started"
                status_cell.fill = danger_fill
                status_cell.font = danger_font
            col += 1
            
            # Progress %
            ws_staff.cell(row=row, column=col, value=progress.completion_percentage if progress else 0)
            col += 1
            
            # Started date
            if progress:
                ws_staff.cell(row=row, column=col, value=progress.started_at.strftime('%d %b %Y'))
            else:
                ws_staff.cell(row=row, column=col, value="-")
            col += 1
            
            # Certificate
            cert_cell = ws_staff.cell(row=row, column=col)
            if certificate:
                cert_cell.value = certificate.certificate_number
                cert_cell.fill = success_fill
            else:
                cert_cell.value = "No"
            col += 1
        
        row += 1
    
    # ==========================================
    # Sheet 3: Not Enrolled Staff
    # ==========================================
    ws_not_enrolled = wb.create_sheet("Not Started")
    
    not_enrolled_headers = ["#", "Full Name", "Email (AD Account)", "Last Login", "Login Status", "Action Required"]
    for col, header in enumerate(not_enrolled_headers, 1):
        cell = ws_not_enrolled.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    row = 2
    idx = 1
    for banker in all_bankers.order_by('first_name', 'last_name'):
        # Check if user has started any course
        has_progress = InteractiveCourseProgress.objects.filter(user=banker).exists()
        if not has_progress:
            ws_not_enrolled.cell(row=row, column=1, value=idx)
            ws_not_enrolled.cell(row=row, column=2, value=banker.get_full_name() or banker.username)
            ws_not_enrolled.cell(row=row, column=3, value=banker.email)
            
            if banker.last_login:
                ws_not_enrolled.cell(row=row, column=4, value=banker.last_login.strftime('%d %b %Y %H:%M'))
                login_cell = ws_not_enrolled.cell(row=row, column=5, value="Logged In")
                login_cell.fill = warning_fill
                ws_not_enrolled.cell(row=row, column=6, value="Has logged in but not started course")
            else:
                ws_not_enrolled.cell(row=row, column=4, value="Never")
                login_cell = ws_not_enrolled.cell(row=row, column=5, value="Never Logged In")
                login_cell.fill = danger_fill
                login_cell.font = danger_font
                ws_not_enrolled.cell(row=row, column=6, value="Needs to login and start course")
            
            row += 1
            idx += 1
    
    # ==========================================
    # Sheet 4: Certified Staff
    # ==========================================
    ws_certified = wb.create_sheet("Certified Staff")
    
    certified_headers = ["#", "Full Name", "Email", "Course", "Certificate Number", "Issue Date", "Score"]
    for col, header in enumerate(certified_headers, 1):
        cell = ws_certified.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    certificates = Certificate.objects.filter(is_valid=True).select_related('user', 'interactive_course').order_by('-issue_date')
    
    for row, cert in enumerate(certificates, 2):
        ws_certified.cell(row=row, column=1, value=row-1)
        ws_certified.cell(row=row, column=2, value=cert.user.get_full_name())
        ws_certified.cell(row=row, column=3, value=cert.user.email)
        ws_certified.cell(row=row, column=4, value=cert.interactive_course.title if cert.interactive_course else cert.course.title if cert.course else "N/A")
        ws_certified.cell(row=row, column=5, value=cert.certificate_number)
        ws_certified.cell(row=row, column=6, value=cert.issue_date.strftime('%d %b %Y') if cert.issue_date else "N/A")
        ws_certified.cell(row=row, column=7, value=f"{cert.overall_score}%" if cert.overall_score else "N/A")
    
    # ==========================================
    # Sheet 5: Legacy Course Overview (Original)
    # ==========================================
    ws_overview = wb.create_sheet("Course Overview")
    
    overview_headers = [
        "Course ID", "Course Title", "Created By", "Created Date", 
        "Total Enrollments", "Completed", "In Progress", "Completion Rate (%)",
        "Avg Quiz Score", "Videos Count", "Questions Count", "Is Published"
    ]
    
    for col, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get course data
    courses = Course.objects.all().order_by('-created_at')
    
    for row, course in enumerate(courses, 2):
        enrollments = course.enrollments.all()
        completed = enrollments.filter(is_completed=True).count()
        total = enrollments.count()
        completion_rate = (completed / total * 100) if total > 0 else 0
        
        # Get average quiz score for this course
        quiz_attempts = QuizAttempt.objects.filter(course=course, completed_at__isnull=False)
        avg_score = quiz_attempts.aggregate(Avg('score'))['score__avg'] or 0
        
        ws_overview.cell(row=row, column=1, value=course.id)
        ws_overview.cell(row=row, column=2, value=course.title)
        ws_overview.cell(row=row, column=3, value=course.created_by.get_full_name() if course.created_by else "N/A")
        ws_overview.cell(row=row, column=4, value=course.created_at.strftime('%Y-%m-%d'))
        ws_overview.cell(row=row, column=5, value=total)
        ws_overview.cell(row=row, column=6, value=completed)
        ws_overview.cell(row=row, column=7, value=total - completed)
        ws_overview.cell(row=row, column=8, value=round(completion_rate, 1))
        ws_overview.cell(row=row, column=9, value=round(avg_score, 1))
        ws_overview.cell(row=row, column=10, value=course.videos.count())
        ws_overview.cell(row=row, column=11, value=course.questions.count())
        ws_overview.cell(row=row, column=12, value="Yes" if course.is_published else "No")
    
    # Auto-adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Course_Enrollment_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def download_user_performance_report(request):
    """Download detailed Excel report of individual user quiz performance"""
    if not request.user.is_risk_admin():
        messages.error(request, 'Unauthorized access.')
        return redirect('content:dashboard')
    
    # Create workbook
    wb = Workbook()
    
    # Quiz Performance Overview Sheet
    ws_overview = wb.active
    ws_overview.title = "Quiz Performance Overview"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    overview_headers = [
        "User ID", "Full Name", "Email", "Course Title", "Quiz Attempts", 
        "Best Score", "Latest Score", "Avg Score", "Questions Answered", 
        "Correct Answers", "Wrong Answers", "Accuracy (%)", "Last Attempt Date"
    ]
    
    for col, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all quiz attempts with user and course data
    attempts_data = []
    users = User.objects.filter(role='banker')
    
    for user in users:
        courses = Course.objects.filter(enrollments__user=user)
        for course in courses:
            user_attempts = QuizAttempt.objects.filter(
                user=user, course=course, completed_at__isnull=False
            ).order_by('-completed_at')
            
            if user_attempts.exists():
                best_score = user_attempts.order_by('-score').first().score
                latest_score = user_attempts.first().score
                avg_score = user_attempts.aggregate(Avg('score'))['score__avg']
                
                # Get answer statistics
                user_answers = QuizAnswer.objects.filter(
                    attempt__user=user, attempt__course=course
                )
                total_answers = user_answers.count()
                correct_answers = user_answers.filter(is_correct=True).count()
                wrong_answers = total_answers - correct_answers
                accuracy = (correct_answers / total_answers * 100) if total_answers > 0 else 0
                
                attempts_data.append({
                    'user': user,
                    'course': course,
                    'attempts_count': user_attempts.count(),
                    'best_score': best_score,
                    'latest_score': latest_score,
                    'avg_score': avg_score,
                    'total_answers': total_answers,
                    'correct_answers': correct_answers,
                    'wrong_answers': wrong_answers,
                    'accuracy': accuracy,
                    'last_attempt': user_attempts.first().completed_at
                })
    
    # Populate overview sheet
    for row, data in enumerate(attempts_data, 2):
        ws_overview.cell(row=row, column=1, value=data['user'].id)
        ws_overview.cell(row=row, column=2, value=data['user'].get_full_name())
        ws_overview.cell(row=row, column=3, value=data['user'].email)
        ws_overview.cell(row=row, column=4, value=data['course'].title)
        ws_overview.cell(row=row, column=5, value=data['attempts_count'])
        ws_overview.cell(row=row, column=6, value=round(data['best_score'], 1))
        ws_overview.cell(row=row, column=7, value=round(data['latest_score'], 1))
        ws_overview.cell(row=row, column=8, value=round(data['avg_score'], 1))
        ws_overview.cell(row=row, column=9, value=data['total_answers'])
        ws_overview.cell(row=row, column=10, value=data['correct_answers'])
        ws_overview.cell(row=row, column=11, value=data['wrong_answers'])
        ws_overview.cell(row=row, column=12, value=round(data['accuracy'], 1))
        ws_overview.cell(row=row, column=13, value=data['last_attempt'].strftime('%Y-%m-%d %H:%M'))
    
    # Detailed Answers Sheet
    ws_details = wb.create_sheet("Detailed Answer Analysis")
    
    detail_headers = [
        "User Name", "Email", "Course", "Question Text", "Correct Answer", 
        "User Answer", "Is Correct", "Question Topic", "Difficulty", 
        "Points", "Attempt Date", "Attempt Score"
    ]
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get detailed answer data
    user_answers = QuizAnswer.objects.select_related(
        'attempt__user', 'attempt__course', 'question'
    ).order_by('-attempt__completed_at')
    
    for row, answer in enumerate(user_answers, 2):
        attempt = answer.attempt
        question = answer.question
        
        # Find correct answer
        correct_option = question.options.filter(is_correct=True).first()
        correct_text = correct_option.option_text if correct_option else "N/A"
        
        # Get user's selected answers (could be multiple for multiple answer questions)
        selected_options = answer.selected_options.all()
        user_answer_text = ", ".join([opt.option_text for opt in selected_options]) if selected_options else "No Answer"
        
        ws_details.cell(row=row, column=1, value=attempt.user.get_full_name())
        ws_details.cell(row=row, column=2, value=attempt.user.email)
        ws_details.cell(row=row, column=3, value=attempt.course.title)
        ws_details.cell(row=row, column=4, value=question.question_text[:100] + "..." if len(question.question_text) > 100 else question.question_text)
        ws_details.cell(row=row, column=5, value=correct_text)
        ws_details.cell(row=row, column=6, value=user_answer_text)
        ws_details.cell(row=row, column=7, value="✓ Correct" if answer.is_correct else "✗ Wrong")
        ws_details.cell(row=row, column=8, value=question.topic or "General")
        ws_details.cell(row=row, column=9, value=question.get_difficulty_display())
        ws_details.cell(row=row, column=10, value=question.points)
        ws_details.cell(row=row, column=11, value=attempt.completed_at.strftime('%Y-%m-%d %H:%M'))
        ws_details.cell(row=row, column=12, value=round(attempt.score, 1))
        
        # Color code correct/wrong answers
        if answer.is_correct:
            ws_details.cell(row=row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws_details.cell(row=row, column=7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    for ws in [ws_overview, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"User_Performance_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# =====================================================
# Interactive Course Question Management Views
# =====================================================

@login_required
def interactive_question_bank(request, interactive_id):
    """View and manage questions for an interactive course"""
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    
    # Check permission - only course creator or admin can manage questions
    if not request.user.is_risk_admin() and interactive_course.course.created_by != request.user:
        messages.error(request, "You don't have permission to manage questions for this course.")
        return redirect('courses:dashboard')
    
    questions = Question.objects.filter(interactive_course=interactive_course).order_by('topic', '-created_at')
    
    # Get unique topics for filtering
    topics = questions.values_list('topic', flat=True).distinct()
    
    context = {
        'interactive_course': interactive_course,
        'course': interactive_course.course,
        'questions': questions,
        'topics': [t for t in topics if t],
        'question_count': questions.count(),
    }
    return render(request, 'content/interactive_question_bank.html', context)


@login_required
def add_interactive_question(request, interactive_id):
    """Add a new question to an interactive course"""
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    
    # Check permission
    if not request.user.is_risk_admin() and interactive_course.course.created_by != request.user:
        messages.error(request, "You don't have permission to add questions to this course.")
        return redirect('courses:dashboard')
    
    if request.method == 'POST':
        question_text = request.POST.get('question_text', '').strip()
        question_type = request.POST.get('question_type', 'multiple_choice')
        topic = request.POST.get('topic', '').strip()
        difficulty = request.POST.get('difficulty', 'medium')
        points = request.POST.get('points', 1)
        explanation = request.POST.get('explanation', '').strip()
        
        if not question_text:
            messages.error(request, 'Question text is required.')
            return redirect('content:interactive_question_bank', interactive_id=interactive_id)
        
        # Create question
        question = Question.objects.create(
            interactive_course=interactive_course,
            question_text=question_text,
            question_type=question_type,
            topic=topic,
            difficulty=difficulty,
            points=int(points),
            explanation=explanation
        )
        
        # Add options - parse individual option_X and correct_X fields
        option_count = int(request.POST.get('option_count', 4))
        order_index = 0
        
        for i in range(1, option_count + 10):  # Check up to option_count + buffer
            option_text = request.POST.get(f'option_{i}', '').strip()
            if option_text:
                is_correct = request.POST.get(f'correct_{i}') == 'on'
                QuestionOption.objects.create(
                    question=question,
                    option_text=option_text,
                    is_correct=is_correct,
                    order_index=order_index
                )
                order_index += 1
        
        messages.success(request, f'Question added successfully! Total questions: {Question.objects.filter(interactive_course=interactive_course).count()}')
        return redirect('content:interactive_question_bank', interactive_id=interactive_id)
    
    # GET request - show add form
    context = {
        'interactive_course': interactive_course,
        'course': interactive_course.course,
        'action': 'add',
    }
    return render(request, 'content/interactive_question_form.html', context)


@login_required
def edit_interactive_question(request, interactive_id, question_id):
    """Edit an existing question"""
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    question = get_object_or_404(Question, id=question_id, interactive_course=interactive_course)
    
    # Check permission
    if not request.user.is_risk_admin() and interactive_course.course.created_by != request.user:
        messages.error(request, "You don't have permission to edit questions for this course.")
        return redirect('courses:dashboard')
    
    if request.method == 'POST':
        question.question_text = request.POST.get('question_text', '').strip()
        question.question_type = request.POST.get('question_type', 'multiple_choice')
        question.topic = request.POST.get('topic', '').strip()
        question.difficulty = request.POST.get('difficulty', 'medium')
        question.points = int(request.POST.get('points', 1))
        question.explanation = request.POST.get('explanation', '').strip()
        question.save()
        
        # Update options - delete old ones and create new
        question.options.all().delete()
        
        # Parse individual option_X and correct_X fields
        option_count = int(request.POST.get('option_count', 4))
        order_index = 0
        
        for i in range(1, option_count + 10):  # Check up to option_count + buffer
            option_text = request.POST.get(f'option_{i}', '').strip()
            if option_text:
                is_correct = request.POST.get(f'correct_{i}') == 'on'
                QuestionOption.objects.create(
                    question=question,
                    option_text=option_text,
                    is_correct=is_correct,
                    order_index=order_index
                )
                order_index += 1
        
        messages.success(request, 'Question updated successfully!')
        return redirect('content:interactive_question_bank', interactive_id=interactive_id)
    
    # GET request - show edit form
    context = {
        'interactive_course': interactive_course,
        'course': interactive_course.course,
        'question': question,
        'action': 'edit',
    }
    return render(request, 'content/interactive_question_form.html', context)


@login_required
def delete_interactive_question(request, interactive_id, question_id):
    """Delete a question from an interactive course"""
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    question = get_object_or_404(Question, id=question_id, interactive_course=interactive_course)
    
    # Check permission
    if not request.user.is_risk_admin() and interactive_course.course.created_by != request.user:
        messages.error(request, "You don't have permission to delete questions from this course.")
        return redirect('courses:dashboard')
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
    
    return redirect('content:interactive_question_bank', interactive_id=interactive_id)


@login_required
def delete_interactive_course(request, interactive_id):
    """Delete an interactive course and its associated files"""
    from videos.models import InteractiveCourse, InteractiveCourseProgress
    import shutil
    
    interactive_course = get_object_or_404(InteractiveCourse, id=interactive_id)
    
    # Check permission - only risk admins or course creator can delete
    if not request.user.is_risk_admin() and interactive_course.course.created_by != request.user:
        messages.error(request, "You don't have permission to delete this interactive course.")
        return redirect('content:interactive_list')
    
    if request.method == 'POST':
        title = interactive_course.title
        
        # Delete progress records
        InteractiveCourseProgress.objects.filter(interactive_course=interactive_course).delete()
        
        # Delete questions associated with this interactive course
        Question.objects.filter(interactive_course=interactive_course).delete()
        
        # Delete extracted files if they exist
        if interactive_course.extracted_path:
            extracted_dir = os.path.join(settings.MEDIA_ROOT, interactive_course.extracted_path)
            if os.path.exists(extracted_dir):
                try:
                    shutil.rmtree(extracted_dir)
                except Exception as e:
                    pass  # Continue even if file deletion fails
        
        # Delete the package file
        if interactive_course.package_file:
            try:
                interactive_course.package_file.delete(save=False)
            except Exception:
                pass
        
        # Delete the interactive course record
        interactive_course.delete()
        
        messages.success(request, f'Interactive course "{title}" and all associated data have been deleted.')
    
    return redirect('content:interactive_list')
